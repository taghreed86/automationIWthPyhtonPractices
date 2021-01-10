import openpyxl


class TestData:

    def getTestData(self):
        Dict_Data1 = {}
        Dict_Data2 = {}
        workbook = openpyxl.load_workbook("/home/taghreed/PycharmProjects/PageObjectModelTesting/testDataFile.xlsx")
        sheet = workbook.active
        for i in range(1, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                Dict_Data1[sheet.cell(row=1, column=j).value] = sheet.cell(row=2, column=j).value
                Dict_Data2[sheet.cell(row=1, column=j).value] = sheet.cell(row=3, column=j).value

        #print(Dict_Data1)
        #print(Dict_Data2)
        homeTestData = [Dict_Data1, Dict_Data2]
        return homeTestData


